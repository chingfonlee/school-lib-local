from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.services.export_service import _extend_data_rows


def test_extend_data_rows_copies_template_row_style_and_height():
    wb = Workbook()
    ws = wb.active
    template_row = 55
    insert_at = 56

    ws.row_dimensions[template_row].height = 19.5
    thin = Side(style="thin", color="000000")
    for col_idx in range(1, 13):
        cell = ws.cell(template_row, col_idx)
        cell.font = Font(name="Arial", bold=True)
        cell.fill = PatternFill("solid", fgColor="FFEFEFEF")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.number_format = "@"

    _extend_data_rows(ws, template_row=template_row, insert_at=insert_at, extra_rows=3)

    for row in range(56, 59):
        assert ws.row_dimensions[row].height == 19.5
        for col_idx in range(1, 13):
            source = ws.cell(template_row, col_idx)
            target = ws.cell(row, col_idx)
            assert target.font.name == source.font.name
            assert target.font.bold == source.font.bold
            assert target.fill.fgColor.rgb == source.fill.fgColor.rgb
            assert target.border.left.style == source.border.left.style
            assert target.border.right.style == source.border.right.style
            assert target.border.top.style == source.border.top.style
            assert target.border.bottom.style == source.border.bottom.style
            assert target.alignment.horizontal == source.alignment.horizontal
            assert target.alignment.vertical == source.alignment.vertical
            assert target.alignment.wrap_text == source.alignment.wrap_text
            assert target.number_format == source.number_format


def test_extend_data_rows_unmerges_ranges_that_overlap_inserted_data_rows():
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("A57:L57")
    ws.merge_cells("A58:L58")
    ws.merge_cells("A80:L80")

    _extend_data_rows(ws, template_row=55, insert_at=56, extra_rows=3)

    merged_ranges = {str(r) for r in ws.merged_cells.ranges}
    # Rows 57 and 58 were in the inserted area [56, 58] → removed (not shifted)
    assert "A57:L57" not in merged_ranges
    assert "A58:L58" not in merged_ranges
    # A80 is a footer range (min_row=80 >= insert_at=56), shifted +3 → A83
    assert "A83:L83" in merged_ranges
    assert "A80:L80" not in merged_ranges


# ── New tests for footer merged range save/restore ──


def test_footer_merged_ranges_preserved_and_shifted():
    """備註列 + 簽核列 merged ranges 在插列後正確位移，不被刪除。"""
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("B57:L57")  # 備註列：min_row=57 >= insert_at=56
    ws.merge_cells("A58:L58")  # 簽核列：min_row=58 >= insert_at=56

    _extend_data_rows(ws, template_row=55, insert_at=56, extra_rows=5)

    merged_ranges = {str(r) for r in ws.merged_cells.ranges}
    assert "B62:L62" in merged_ranges   # 57 + 5 = 62
    assert "A63:L63" in merged_ranges   # 58 + 5 = 63
    assert "B57:L57" not in merged_ranges
    assert "A58:L58" not in merged_ranges


def test_footer_multiple_merged_ranges_all_shifted():
    """多個 footer 合併範圍（含國小/國中兩種結構）全部正確位移。"""
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("B57:L57")  # 備註（國小格式）
    ws.merge_cells("A58:L58")  # 簽核
    ws.merge_cells("B59:G59")  # 其他 footer 元素

    _extend_data_rows(ws, template_row=55, insert_at=56, extra_rows=10)

    merged_ranges = {str(r) for r in ws.merged_cells.ranges}
    assert "B67:L67" in merged_ranges   # 57 + 10 = 67
    assert "A68:L68" in merged_ranges   # 58 + 10 = 68
    assert "B69:G69" in merged_ranges   # 59 + 10 = 69
    # Originals gone
    assert "B57:L57" not in merged_ranges
    assert "A58:L58" not in merged_ranges
    assert "B59:G59" not in merged_ranges


def test_footer_merged_ranges_unchanged_when_no_extra_rows():
    """extra_rows=0 時函式提早返回，footer 合併範圍不變。"""
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("A57:L57")

    _extend_data_rows(ws, template_row=55, insert_at=56, extra_rows=0)

    merged_ranges = {str(r) for r in ws.merged_cells.ranges}
    assert "A57:L57" in merged_ranges


def test_cross_boundary_merged_range_removed():
    """跨越插入點的 merged range（min_row<insert_at<=max_row）保守移除，不重建。"""
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("A50:L60")  # min_row=50 < 56, max_row=60 >= 56

    _extend_data_rows(ws, template_row=55, insert_at=56, extra_rows=3)

    merged_ranges = {str(r) for r in ws.merged_cells.ranges}
    assert "A50:L60" not in merged_ranges   # 原始位置已移除
    assert "A53:L63" not in merged_ranges   # 不應被重建（非 footer 範圍）


def test_footer_row_height_shifted_after_extension():
    """確認 openpyxl insert_rows 是否自動位移 footer row_dimensions。

    若通過：openpyxl 已自動處理，無需補 step 7。
    若失敗：需在 _extend_data_rows 加入明確 row_dimensions 位移邏輯（plan step 7）。
    """
    wb = Workbook()
    ws = wb.active
    ws.row_dimensions[57].height = 20.0
    ws.row_dimensions[58].height = 22.0

    _extend_data_rows(ws, template_row=55, insert_at=56, extra_rows=5)

    assert ws.row_dimensions[62].height == 20.0, (
        "row 57 height should shift to row 62 after inserting 5 rows at 56"
    )
    assert ws.row_dimensions[63].height == 22.0, (
        "row 58 height should shift to row 63 after inserting 5 rows at 56"
    )
