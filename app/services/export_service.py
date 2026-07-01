import json
import shutil
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string

from app.database import get_connection

# H欄合法值（來源：範本 N5:N13）— 與 validation_service.py 中的常數保持同步
GENERAL_BOOKS_H_ALLOWED = {
    "喜閱網",
    "文化部中小學生優良課外讀物選介",
    "好書大家讀",
    "金鼎獎",
    "文化部Books from Taiwan等具高公信力推薦價值之圖書",
    "國民中小學新生閱讀推動活動入選書單",
    "圖書分級推薦書目、臺灣歷史文化分級推薦書目",
    "其他國內外具公信力單位辦理之獎項(請備註獎項名稱)",
    "學校自選(請備註原因)",
}

GENERAL_BOOKS_H_REQUIRES_NOTES = {
    "其他國內外具公信力單位辦理之獎項(請備註獎項名稱)",
    "學校自選(請備註原因)",
}


@dataclass
class ExportSettings:
    project_id: int
    school_name: str
    approved_budget: float | None
    price_field: str
    subtotal_mode: str
    output_dir: str
    exported_by: int


def _load_export_template_for_project(project_id: int, conn) -> dict:
    """Look up the export template for a project by its export_template_type."""
    project = conn.execute(
        "SELECT project_type, export_template_type FROM procurement_projects WHERE id = ?",
        (project_id,),
    ).fetchone()
    if project is None:
        raise ValueError("採購專案不存在")
    tmpl = conn.execute(
        "SELECT * FROM export_templates WHERE project_type = ? ORDER BY id LIMIT 1",
        (project["export_template_type"],),
    ).fetchone()
    if tmpl is None:
        raise ValueError(
            f"找不到匯出範本（project_type={project['export_template_type']}）"
            "，請至導覽列「範本管理」確認範本已設定。"
        )
    result = dict(tmpl)
    result["_project_type"] = project["project_type"]
    return result


def _resolve_field(book: dict, field: str) -> str:
    overrides = json.loads(book.get("user_overrides") or "{}")
    if field in overrides and overrides[field] not in (None, ""):
        return str(overrides[field])
    v = book.get(field)
    if v not in (None, ""):
        return str(v)
    return ""


def _get_price(book: dict, price_field: str) -> float:
    overrides = json.loads(book.get("user_overrides") or "{}")
    val = overrides.get(price_field) or book.get(price_field)
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def _resolve_eligibility_for_export(book: dict, project_type: str) -> str:
    value = _resolve_field(book, "eligibility_label")
    if project_type == "general_books_jh" and value == "推薦":
        return "必選"
    return value


def _is_force_owned(book: dict) -> bool:
    ov = json.loads(book.get("user_overrides") or "{}")
    return ov.get("force_owned") is True


def export_local_culture(settings: ExportSettings) -> str:
    conn = get_connection()
    tmpl = _load_export_template_for_project(settings.project_id, conn)
    conn.close()

    col_map = json.loads(tmpl["column_mappings"])

    def col(key: str) -> int:
        return column_index_from_string(col_map[key])

    template_path = tmpl["template_file_path"]
    data_start = tmpl["data_start_row"]
    max_rows = tmpl["max_rows"]
    example_row = data_start - 1
    template_end = data_start + max_rows - 1
    summary_row = data_start + max_rows

    Path(settings.output_dir).mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    out_filename = f"本土文化採購書單_{ts}.xlsx"
    out_path = str(Path(settings.output_dir) / out_filename)

    shutil.copy2(template_path, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    school_name_prefix = "校名："
    budget_prefix = "本案核定金額：新臺幣"
    budget_suffix = "元"

    ws[tmpl["school_name_cell"]] = school_name_prefix + settings.school_name
    if settings.approved_budget is not None:
        ws[tmpl["approved_budget_cell"]] = (
            budget_prefix + f"{settings.approved_budget:,.0f}" + budget_suffix
        )

    if example_row >= 1:
        for key in col_map:
            ws.cell(row=example_row, column=col(key)).value = None

    conn = get_connection()
    books = conn.execute(
        "SELECT si.*, "
        "COALESCE("
        "  (SELECT bm.match_status FROM book_matches bm "
        "   WHERE bm.vendor_book_id = si.vendor_book_id "
        "     AND bm.match_status != 'same_title_different_isbn' "
        "   ORDER BY bm.id DESC LIMIT 1), "
        "  si.match_status_at_selection, 'available'"
        ") AS match_status "
        "FROM selection_items si "
        "WHERE si.project_id = ? AND si.selected_quantity > 0 "
        "ORDER BY si.vendor_seq, si.id",
        (settings.project_id,),
    ).fetchall()
    conn.close()

    all_books = [dict(b) for b in books]
    exportable = [
        b for b in all_books
        if b["match_status"] in ("available", "missing_isbn", "invalid_isbn")
        or (b["match_status"] == "already_owned" and _is_force_owned(b))
    ]

    extra_rows_needed = max(0, len(exportable) - (template_end - data_start + 1))
    if extra_rows_needed > 0:
        ws.insert_rows(summary_row, extra_rows_needed)
        summary_row += extra_rows_needed

    total_qty = 0
    total_amount = 0.0

    for i, book in enumerate(exportable):
        row = data_start + i
        if row > template_end + extra_rows_needed:
            break

        quantity = book["selected_quantity"]
        price = _get_price(book, settings.price_field)

        if settings.subtotal_mode == "quantity_times_list_price":
            subtotal = quantity * _get_price(book, "list_price")
        else:
            subtotal = quantity * _get_price(book, "purchase_price")

        ws.cell(row=row, column=col("sort_order")).value = i + 1
        ws.cell(row=row, column=col("title")).value = _resolve_field(book, "title")
        ws.cell(row=row, column=col("author")).value = _resolve_field(book, "author")
        ws.cell(row=row, column=col("publisher")).value = _resolve_field(book, "publisher")
        ws.cell(row=row, column=col("isbn")).value = (
            _resolve_field(book, "isbn_normalized") or _resolve_field(book, "isbn")
        )
        ws.cell(row=row, column=col("quantity")).value = quantity
        ws.cell(row=row, column=col("price")).value = price if price else None
        ws.cell(row=row, column=col("subtotal")).value = subtotal if subtotal else None
        ws.cell(row=row, column=col("award_item")).value = _resolve_field(book, "award_item")
        ws.cell(row=row, column=col("notes")).value = book.get("notes") or None

        total_qty += quantity
        total_amount += subtotal

    ws.cell(row=summary_row, column=col("quantity")).value = total_qty
    ws.cell(row=summary_row, column=col("subtotal")).value = total_amount

    wb.save(out_path)

    now = datetime.now(timezone.utc).isoformat()
    conn = get_connection()
    row = conn.execute(
        "INSERT INTO export_jobs"
        "(project_id, school_name, approved_budget, price_field, subtotal_mode, "
        "template_path, output_filename, output_path, exported_by, exported_at, "
        "record_count, total_amount, export_template_id) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (
            settings.project_id,
            settings.school_name,
            settings.approved_budget,
            settings.price_field,
            settings.subtotal_mode,
            template_path,
            out_filename,
            out_path,
            settings.exported_by,
            now,
            len(exportable),
            total_amount,
            tmpl["id"],
        ),
    )
    job_id = row.lastrowid
    conn.commit()
    conn.close()

    return str(job_id)


def export_general_books(settings: ExportSettings) -> str:
    conn = get_connection()
    tmpl = _load_export_template_for_project(settings.project_id, conn)
    conn.close()

    col_map = json.loads(tmpl["column_mappings"])

    def col(key: str) -> int:
        return column_index_from_string(col_map[key])

    template_path = tmpl["template_file_path"]
    project_type = tmpl.get("_project_type") or ""
    sheet_name = tmpl.get("sheet_name") or ""
    data_start = tmpl["data_start_row"]
    max_rows = tmpl["max_rows"]
    example_row = data_start - 1
    template_end = data_start + max_rows - 1
    summary_row = data_start + max_rows

    Path(settings.output_dir).mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    out_filename = f"一般圖書採購書單_{ts}.xlsx"
    out_path = str(Path(settings.output_dir) / out_filename)

    shutil.copy2(template_path, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    school_name_cell = tmpl.get("school_name_cell") or ""
    approved_budget_cell = tmpl.get("approved_budget_cell") or ""
    school_name_prefix = "校名："
    budget_prefix = "本案核定金額：新臺幣"
    budget_suffix = "元"

    if school_name_cell:
        ws[school_name_cell] = school_name_prefix + settings.school_name
    if approved_budget_cell and settings.approved_budget is not None:
        ws[approved_budget_cell] = (
            budget_prefix + f"{settings.approved_budget:,.0f}" + budget_suffix
        )

    if example_row >= 1:
        for key in col_map:
            ws.cell(row=example_row, column=col(key)).value = None

    conn = get_connection()
    books = conn.execute(
        "SELECT si.*, "
        "COALESCE("
        "  (SELECT bm.match_status FROM book_matches bm "
        "   WHERE bm.vendor_book_id = si.vendor_book_id "
        "     AND bm.match_status != 'same_title_different_isbn' "
        "   ORDER BY bm.id DESC LIMIT 1), "
        "  si.match_status_at_selection, 'available'"
        ") AS match_status "
        "FROM selection_items si "
        "WHERE si.project_id = ? AND si.selected_quantity > 0 "
        "ORDER BY si.vendor_seq, si.id",
        (settings.project_id,),
    ).fetchall()
    conn.close()

    all_books = [dict(b) for b in books]
    exportable = [
        b for b in all_books
        if b["match_status"] in ("available", "missing_isbn", "invalid_isbn")
        or (b["match_status"] == "already_owned" and _is_force_owned(b))
    ]

    extra_rows_needed = max(0, len(exportable) - (template_end - data_start + 1))
    if extra_rows_needed > 0:
        ws.insert_rows(summary_row, extra_rows_needed)
        summary_row += extra_rows_needed

    total_qty = 0
    total_amount = 0.0

    for i, book in enumerate(exportable):
        row = data_start + i
        if row > template_end + extra_rows_needed:
            break

        quantity = book["selected_quantity"]
        price = _get_price(book, settings.price_field)
        subtotal = quantity * price

        ws.cell(row=row, column=col("eligibility_label")).value = _resolve_eligibility_for_export(book, project_type) or None
        ws.cell(row=row, column=col("sort_order")).value = i + 1
        ws.cell(row=row, column=col("title")).value = _resolve_field(book, "title")
        ws.cell(row=row, column=col("author")).value = _resolve_field(book, "author") or None
        ws.cell(row=row, column=col("publisher")).value = _resolve_field(book, "publisher") or None
        ws.cell(row=row, column=col("isbn")).value = (
            _resolve_field(book, "isbn_normalized") or _resolve_field(book, "isbn") or None
        )
        ws.cell(row=row, column=col("quantity")).value = quantity
        if "recommendation_source" in col_map:
            ws.cell(row=row, column=col("recommendation_source")).value = _resolve_field(book, "recommendation_source") or None
        if "policy_topic" in col_map:
            ws.cell(row=row, column=col("policy_topic")).value = _resolve_field(book, "policy_topic") or None
        ws.cell(row=row, column=col("price")).value = price if price else None
        ws.cell(row=row, column=col("subtotal")).value = subtotal if subtotal else None
        if "award_notes" in col_map:
            ws.cell(row=row, column=col("award_notes")).value = _resolve_field(book, "award_notes") or None
        if "notes" in col_map:
            ws.cell(row=row, column=col("notes")).value = book.get("notes") or None

        total_qty += quantity
        total_amount += subtotal

    ws.cell(row=summary_row, column=col("quantity")).value = total_qty
    ws.cell(row=summary_row, column=col("subtotal")).value = total_amount

    wb.save(out_path)

    now = datetime.now(timezone.utc).isoformat()
    conn = get_connection()
    row_result = conn.execute(
        "INSERT INTO export_jobs"
        "(project_id, school_name, approved_budget, price_field, subtotal_mode, "
        "template_path, output_filename, output_path, exported_by, exported_at, "
        "record_count, total_amount, export_template_id) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (
            settings.project_id,
            settings.school_name,
            settings.approved_budget,
            settings.price_field,
            settings.subtotal_mode,
            template_path,
            out_filename,
            out_path,
            settings.exported_by,
            now,
            len(exportable),
            total_amount,
            tmpl["id"],
        ),
    )
    job_id = row_result.lastrowid
    conn.commit()
    conn.close()

    return str(job_id)
