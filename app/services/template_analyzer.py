"""
Excel 匯出範本自動偵測服務。
從範本的標題列推斷欄位位置，並偵測學校名稱/預算儲存格。
"""
import openpyxl
from openpyxl.utils import get_column_letter

# 各專案類型的預期欄位 + 偵測關鍵字（由長到短，越長越優先）
FIELD_KEYWORDS: dict[str, dict[str, list[str]]] = {
    "local_culture": {
        "sort_order": ["排序"],
        "title":      ["書名"],
        "author":     ["作者"],
        "publisher":  ["出版社", "出版者"],
        "isbn":       ["ISBN"],
        "quantity":   ["採購數量", "數量"],
        "price":      ["定價"],
        "subtotal":   ["小計"],
        "award_item": ["獲獎項目"],
        "notes":      ["備註"],
    },
    "local_culture_jh": {
        "sort_order": ["排序"],
        "title":      ["書名"],
        "author":     ["作者"],
        "publisher":  ["出版社", "出版者"],
        "isbn":       ["ISBN"],
        "quantity":   ["採購數量", "數量"],
        "price":      ["定價"],
        "subtotal":   ["小計"],
        "award_item": ["獲獎項目"],
        "notes":      ["備註"],
    },
    "general_books": {
        "eligibility_label":     ["必選/自選", "必選或自選"],
        "sort_order":            ["排序"],
        "title":                 ["書名"],
        "author":                ["作者"],
        "publisher":             ["出版社", "出版者"],
        "isbn":                  ["ISBN"],
        "quantity":              ["採購數量", "數量"],
        "recommendation_source": ["推薦來源"],
        "policy_topic":          ["重要政策議題", "政策議題", "重要政策", "議題"],
        "price":                 ["定價"],
        "subtotal":              ["小計"],
        "award_notes":           ["獲獎備註", "備註"],
    },
    "general_books_jh": {
        "sort_order":        ["排序"],
        "eligibility_label": ["必選/自選", "必選或自選"],
        "policy_topic":      ["重要政策議題", "政策議題", "重要政策", "議題"],
        "title":             ["書名"],
        "author":            ["作者"],
        "publisher":         ["出版社", "出版者"],
        "isbn":              ["ISBN"],
        "quantity":          ["採購數量", "數量"],
        "price":             ["定價"],
        "subtotal":          ["小計"],
        "recommendation_source": ["獲獎項目"],
        "notes":             ["備註"],
    },
}

FIELD_LABELS: dict[str, str] = {
    "sort_order":            "排序",
    "title":                 "書名",
    "author":                "作者",
    "publisher":             "出版社",
    "isbn":                  "ISBN",
    "quantity":              "採購數量",
    "price":                 "定價",
    "subtotal":              "小計",
    "award_item":            "獲獎項目（本土文化）",
    "award_notes":           "獲獎項目／備註（一般圖書）",
    "notes":                 "備註",
    "eligibility_label":     "必選／自選",
    "recommendation_source": "推薦來源",
    "policy_topic":          "政策議題",
}

_ALL_SCAN_KEYWORDS = ["書名", "作者", "ISBN", "排序", "數量", "定價", "小計", "備註", "出版"]


def _matches(cell_text: str, keywords: list[str]) -> bool:
    ct = cell_text.strip().upper()
    for kw in sorted(keywords, key=lambda k: -len(k)):
        if kw.upper() in ct:
            return True
    return False


def _find_header_row(ws, max_scan: int = 12) -> int:
    best_row, best_score = 4, 0
    for r in range(1, min(max_scan + 1, ws.max_row + 1)):
        score = 0
        for cell in ws[r]:
            if not cell.value or not isinstance(cell.value, str):
                continue
            if _matches(cell.value, _ALL_SCAN_KEYWORDS):
                score += 1
        if score > best_score:
            best_score, best_row = score, r
    return best_row


def _find_special_cells(ws, header_row: int) -> tuple[str, str]:
    school_cell = ""
    budget_cell = ""
    for r in range(1, header_row):
        for cell in ws[r]:
            if not cell.value or not isinstance(cell.value, str):
                continue
            text = cell.value.strip()
            if "校名" in text and not school_cell:
                school_cell = cell.coordinate
            if "核定金額" in text and not budget_cell:
                budget_cell = cell.coordinate
    return school_cell, budget_cell


def _find_data_start_row(ws, header_row: int) -> int:
    next_row = header_row + 1
    if next_row <= ws.max_row:
        for cell in ws[next_row]:
            if cell.value and isinstance(cell.value, str) and "範例" in cell.value:
                return next_row + 1
    return header_row + 1


def analyze_template(file_path: str, project_type: str) -> dict:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    header_row = _find_header_row(ws)
    school_cell, budget_cell = _find_special_cells(ws, header_row)
    data_start_row = _find_data_start_row(ws, header_row)

    # 讀取標題列所有非空儲存格
    all_headers = []
    for cell in ws[header_row]:
        if cell.value is not None:
            text = str(cell.value).strip()
            if text:
                all_headers.append({
                    "col": get_column_letter(cell.column),
                    "text": text,
                })

    # 依關鍵字比對欄位
    keywords = FIELD_KEYWORDS.get(project_type, {})
    detected: dict[str, str] = {}
    used_cols: set[str] = set()

    for field, kws in keywords.items():
        for hc in all_headers:
            if hc["col"] in used_cols:
                continue
            if _matches(hc["text"], kws):
                detected[field] = hc["col"]
                used_cols.add(hc["col"])
                break

    missing = [f for f in keywords if f not in detected]
    max_rows = max(50, ws.max_row - data_start_row - 2)

    return {
        "sheet_name":          ws.title,
        "header_row":          header_row,
        "data_start_row":      data_start_row,
        "max_rows":            min(max_rows, 200),
        "school_name_cell":    school_cell,
        "approved_budget_cell": budget_cell,
        "column_mappings":     detected,
        "missing_fields":      missing,
        "all_headers":         all_headers,
        "field_labels":        FIELD_LABELS,
    }
