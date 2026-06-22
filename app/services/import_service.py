import json
import io
import math
import re
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd

from app.database import get_connection
from app.services.isbn_service import normalize_isbn, get_isbn_status
from app.services.completeness_service import compute as compute_completeness
from app.services.match_service import run_match

LIBRARY_COLUMN_HINTS = {
    "isbn": ["b02", "isbn", "條碼", "isbn碼", "isbn號"],
    "title": ["b03", "書名", "title"],
    "author": ["b07", "作者", "author"],
    "publisher": ["b10", "出版社", "publisher"],
    "publish_year": ["b12", "出版年", "出版年份", "publish_year"],
    "price": ["b16", "價格", "定價", "price"],
    "library_record_id": ["b04", "書目識別號", "書目編號"],
}

VENDOR_COLUMN_HINTS = {
    "award_item": ["獲獎項目", "獎項", "award"],
    "vendor_seq": ["序號", "編號", "no", "no."],
    "title": ["書名", "title"],
    "author": ["作者", "author"],
    "isbn": ["條碼", "isbn", "isbn碼", "isbn號", "條形碼"],
    "publish_date": ["出版日期", "出版年", "publish_date"],
    "list_price": ["定價", "list_price"],
    "purchase_price": ["單價", "採購價", "purchase_price"],
    "publisher": ["出版社", "publisher"],
    "age_range": ["適合年齡", "年齡", "age_range"],
    "category": ["分類", "category"],
    "book_type": ["類型", "書本類型", "book_type"],
    "summary": ["summary_80_120", "摘要", "summary"],
    "source_url": ["連結", "url", "link", "source_url"],
    "recommendation_source": ["award_template", "推薦來源", "recommendation_source"],
    "eligibility_label": ["eligible_label", "資格標籤", "必選推薦", "eligibility_label"],
    "policy_topic": ["topic", "議題", "policy_topic"],
    "classification_number": ["CIP", "分類號", "圖書分類號", "類號"],
    "award_notes": ["award_notes", "備註", "notes"],
}


def _normalize_header(value: str) -> str:
    return "".join(str(value).strip().lower().split())


def _match_columns(df_columns: list[str], hints: dict) -> tuple[dict, list[str]]:
    """Returns (mapping {src_col: sys_field}, unmapped_sys_fields)."""
    mapping = {}
    lower_cols = {_normalize_header(c): c for c in df_columns}
    for field, candidates in hints.items():
        for cand in candidates:
            key = _normalize_header(cand)
            if key in lower_cols:
                mapping[lower_cols[key]] = field
                break
    mapped_targets = set(mapping.values())
    unmapped = [field for field in hints if field not in mapped_targets]
    return mapping, unmapped


def _list_excel_sheets(file_bytes: bytes, engine: str) -> list[str]:
    """Return sheet names from an xls/xlsx workbook."""
    if engine == "openpyxl":
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheets = list(wb.sheetnames)
        wb.close()
    else:
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        sheets = wb.sheet_names()
    return sheets


def _detect_header_row(file_bytes: bytes, engine: str, hints: dict, sheet_name=0) -> int:
    """Score first 20 rows and return the best 0-based header row index."""
    preview = pd.read_excel(
        io.BytesIO(file_bytes),
        engine=engine,
        sheet_name=sheet_name,
        header=None,
        nrows=20,
        dtype=str,
    )
    hint_tokens = {
        str(token).strip().lower()
        for candidates in hints.values()
        for token in candidates
    }
    best_row = 0
    best_score = -1
    for idx, row in preview.iterrows():
        values = [str(v).strip().lower() for v in row.tolist() if pd.notna(v)]
        score = sum(1 for value in values if value in hint_tokens)
        if score > best_score:
            best_score = score
            best_row = int(idx)
    return best_row


def _read_excel_with_detected_header(
    file_bytes: bytes,
    engine: str,
    hints: dict,
    sheet_name=0,
) -> pd.DataFrame:
    best_row = _detect_header_row(file_bytes, engine, hints, sheet_name=sheet_name)
    return pd.read_excel(
        io.BytesIO(file_bytes),
        engine=engine,
        sheet_name=sheet_name,
        header=best_row,
        dtype=str,
    )


def _is_blank_or_total_row(values: list) -> bool:
    cleaned = [str(v).strip() for v in values if v not in (None, "") and not pd.isna(v)]
    if not cleaned:
        return True
    return any(value in {"合計", "總計"} for value in cleaned)


def preview_excel(
    file_bytes: bytes,
    filename: str,
    sheet_name: str | None = None,
    header_row: int | None = None,
) -> dict:
    """
    Preview Excel without writing to DB.
    Returns sheet list, guessed header row (0-based), source columns, and suggested field mappings.
    header_row input is 0-based (pandas convention).
    """
    engine = "xlrd" if filename.lower().endswith(".xls") else "openpyxl"

    if engine == "openpyxl":
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheets = list(wb.sheetnames)
        wb.close()
    else:
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        sheets = wb.sheet_names()

    active_sheet = sheet_name if sheet_name else (sheets[0] if sheets else 0)

    if header_row is None:
        guessed_header_row = _detect_header_row(
            file_bytes, engine, VENDOR_COLUMN_HINTS, sheet_name=active_sheet
        )
    else:
        guessed_header_row = header_row

    df = pd.read_excel(
        io.BytesIO(file_bytes),
        engine=engine,
        sheet_name=active_sheet,
        header=guessed_header_row,
        dtype=str,
    )

    source_columns = []
    for c in df.columns:
        cs = str(c).strip()
        if cs and not cs.lower().startswith("unnamed:"):
            source_columns.append(cs)

    src_to_sys, unmapped_sys = _match_columns(source_columns, VENDOR_COLUMN_HINTS)
    suggested_mappings = {v: k for k, v in src_to_sys.items()}

    mapped_source = set(src_to_sys.keys())
    extra_source_columns = [c for c in source_columns if c not in mapped_source]

    return {
        "sheets": sheets,
        "active_sheet": str(active_sheet),
        "guessed_header_row": guessed_header_row,
        "source_columns": source_columns,
        "suggested_mappings": suggested_mappings,
        "unmapped_system_fields": unmapped_sys,
        "extra_source_columns": extra_source_columns,
    }


def confirm_import(
    file_bytes: bytes,
    filename: str,
    project_id: int,
    sheet_name: str | None,
    header_row: int,
    mappings: dict,
    extra_field_settings: list,
    user_id: int,
    profile_id: int | None = None,
) -> dict:
    """
    Formally import vendor books using user-confirmed mappings (sys_field → src_col).
    header_row is 0-based.
    """
    now = datetime.now(timezone.utc).isoformat()
    engine = "xlrd" if filename.lower().endswith(".xls") else "openpyxl"
    sheet = sheet_name if sheet_name else 0

    pp_formula_map: dict = {}
    if engine == "openpyxl":
        pp_src_col = mappings.get("purchase_price")
        if pp_src_col:
            pp_formula_map = _build_formula_fallback(file_bytes, sheet, header_row, pp_src_col)

    df = pd.read_excel(
        io.BytesIO(file_bytes),
        engine=engine,
        sheet_name=sheet,
        header=header_row,
        dtype=str,
    )
    df.columns = [str(c).strip() for c in df.columns]

    conn = get_connection()
    proj_row = conn.execute(
        "SELECT project_type FROM procurement_projects WHERE id = ?", (project_id,)
    ).fetchone()
    proj_type = proj_row["project_type"] if proj_row else None
    _clear_vendor_books_for_project(conn, project_id)
    batch_row = conn.execute(
        "INSERT INTO import_batches(project_id, batch_type, original_filename, profile_id, "
        "imported_by, imported_at) VALUES (?, 'vendor_books', ?, ?, ?, ?)",
        (project_id, filename, profile_id, user_id, now),
    )
    batch_id = batch_row.lastrowid

    records_inserted = 0
    skipped_count = 0

    for enum_idx, (_, row) in enumerate(df.iterrows()):
        raw = {col: (None if pd.isna(row[col]) else str(row[col]).strip()) for col in df.columns}
        source_row_number = header_row + 2 + enum_idx  # 1-based Excel row

        def get_field(sys_field: str):
            src_col = mappings.get(sys_field)
            if src_col:
                v = raw.get(src_col)
                return v if v else None
            return None

        isbn_raw = get_field("isbn")
        title = get_field("title")
        if _is_blank_or_total_row([isbn_raw, title, get_field("author"), get_field("publisher")]):
            skipped_count += 1
            continue

        isbn_norm = normalize_isbn(isbn_raw)
        isbn_status = get_isbn_status(isbn_raw)

        extra_fields = {}
        for col_name in (extra_field_settings or []):
            val = raw.get(col_name)
            if val:
                extra_fields[col_name] = val

        list_price = _to_float(get_field("list_price"))
        purchase_price = _to_float(get_field("purchase_price"))
        if purchase_price is None and list_price is not None:
            formula = pp_formula_map.get(enum_idx)
            if formula:
                purchase_price = _resolve_formula_purchase_price(formula, list_price)

        book = {
            "title": title,
            "list_price": list_price,
            "purchase_price": purchase_price,
            "author": get_field("author"),
            "publisher": get_field("publisher"),
            "award_item": get_field("award_item"),
            "eligibility_label": get_field("eligibility_label"),
            "recommendation_source": get_field("recommendation_source"),
        }
        completeness = compute_completeness(book, project_type=proj_type)

        conn.execute(
            "INSERT INTO vendor_books"
            "(batch_id, award_item, vendor_seq, title, author, isbn, isbn_normalized, "
            "publish_date, list_price, purchase_price, publisher, age_range, "
            "isbn_status, completeness_status, extra_fields, source_row_number, raw_row, "
            "category, book_type, policy_topic, summary, source_url, recommendation_source, "
            "eligibility_label, classification_number, award_notes) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                batch_id,
                book["award_item"],
                get_field("vendor_seq"),
                book["title"],
                book["author"],
                isbn_raw,
                isbn_norm,
                get_field("publish_date"),
                book["list_price"],
                book["purchase_price"],
                book["publisher"],
                get_field("age_range"),
                isbn_status,
                completeness,
                json.dumps(extra_fields, ensure_ascii=False) if extra_fields else None,
                source_row_number,
                json.dumps(raw, ensure_ascii=False),
                get_field("category"),
                get_field("book_type"),
                get_field("policy_topic"),
                get_field("summary"),
                get_field("source_url"),
                get_field("recommendation_source"),
                get_field("eligibility_label"),
                get_field("classification_number"),
                get_field("award_notes"),
            ),
        )
        records_inserted += 1

    conn.execute(
        "UPDATE import_batches SET record_count = ? WHERE id = ?",
        (records_inserted, batch_id),
    )
    conn.commit()
    conn.close()

    result = {
        "batch_id": batch_id,
        "record_count": records_inserted,
        "skipped_count": skipped_count,
    }
    try:
        result["match_stats"] = run_match(project_id)
    except Exception as e:
        result["match_rerun_error"] = str(e)
    return result


def import_library_holdings(
    file_bytes: bytes,
    filename: str,
    user_id: int,
    profile_id: int | None = None,
    column_overrides: dict | None = None,
) -> dict:
    now = datetime.now(timezone.utc).isoformat()
    engine = "xlrd" if filename.lower().endswith(".xls") else "openpyxl"
    sheets = _list_excel_sheets(file_bytes, engine)

    conn = get_connection()
    try:
        _clear_library_holdings(conn)

        batch_row = conn.execute(
            "INSERT INTO import_batches(project_id, batch_type, original_filename, profile_id, "
            "imported_by, imported_at) VALUES (NULL, 'library_holdings', ?, ?, ?, ?)",
            (filename, profile_id, user_id, now),
        )
        batch_id = batch_row.lastrowid

        records_inserted = 0
        sheet_summaries = []
        skipped_sheets = []
        last_mapping: dict = {}
        last_unmapped: list = []

        for sheet_name in sheets:
            result = _read_library_sheet(file_bytes, engine, sheet_name)
            if result is None:
                skipped_sheets.append({"sheet_name": str(sheet_name), "reason": "無 isbn 或 title 欄位"})
                continue
            df, mapping, unmapped = result
            if column_overrides:
                mapping.update(column_overrides)
            n = _insert_library_holding_rows(conn, batch_id, df, mapping)
            records_inserted += n
            sheet_summaries.append({"sheet_name": str(sheet_name), "record_count": n})
            last_mapping, last_unmapped = mapping, unmapped

        if records_inserted == 0:
            raise ValueError("館藏檔案未匯入任何有效資料，請確認 Excel 格式與欄位名稱")

        conn.execute(
            "UPDATE import_batches SET record_count = ? WHERE id = ?",
            (records_inserted, batch_id),
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    result = {
        "batch_id": batch_id,
        "replaced": True,
        "record_count": records_inserted,
        "sheet_summaries": sheet_summaries,
        "skipped_sheets": skipped_sheets,
        "unmapped_fields": last_unmapped,
        "column_mapping": last_mapping,
        "match_rerun": False,
        "affected_projects": [],
        "match_stats_by_project": {},
    }
    try:
        conn2 = get_connection()
        project_ids = [
            r[0]
            for r in conn2.execute(
                "SELECT DISTINCT project_id FROM import_batches "
                "WHERE batch_type='vendor_books' AND project_id IS NOT NULL"
            ).fetchall()
        ]
        conn2.close()
        for pid in project_ids:
            stats = run_match(pid)
            result["affected_projects"].append(pid)
            result["match_stats_by_project"][str(pid)] = stats
        result["match_rerun"] = bool(result["affected_projects"])
    except Exception as e:
        result["match_rerun_error"] = str(e)
    return result


def import_vendor_books(
    file_bytes: bytes,
    filename: str,
    project_id: int,
    user_id: int,
    profile_id: int | None = None,
    column_overrides: dict | None = None,
) -> dict:
    """Legacy single-step import with auto-detected column mapping."""
    now = datetime.now(timezone.utc).isoformat()
    conn = get_connection()
    proj_row = conn.execute(
        "SELECT project_type FROM procurement_projects WHERE id = ?", (project_id,)
    ).fetchone()
    proj_type = proj_row["project_type"] if proj_row else None

    df = _read_excel_with_detected_header(file_bytes, "openpyxl", VENDOR_COLUMN_HINTS)
    df.columns = [str(c).strip() for c in df.columns]
    mapping, unmapped = _match_columns(list(df.columns), VENDOR_COLUMN_HINTS)
    if column_overrides:
        mapping.update(column_overrides)

    _clear_vendor_books_for_project(conn, project_id)
    batch_row = conn.execute(
        "INSERT INTO import_batches(project_id, batch_type, original_filename, profile_id, "
        "imported_by, imported_at) VALUES (?, 'vendor_books', ?, ?, ?, ?)",
        (project_id, filename, profile_id, user_id, now),
    )
    batch_id = batch_row.lastrowid

    records_inserted = 0
    reverse_map = {v: k for k, v in mapping.items()}

    pp_formula_map: dict = {}
    pp_src_col = reverse_map.get("purchase_price")
    if pp_src_col:
        _vb_header_row = _detect_header_row(file_bytes, "openpyxl", VENDOR_COLUMN_HINTS)
        pp_formula_map = _build_formula_fallback(file_bytes, 0, _vb_header_row, pp_src_col)

    for enum_idx, (_, row) in enumerate(df.iterrows()):
        raw = {col: (None if pd.isna(row[col]) else str(row[col]).strip()) for col in df.columns}

        def get_field(field: str):
            src = reverse_map.get(field)
            if src:
                v = raw.get(src)
                return v if v else None
            return None

        isbn_raw = get_field("isbn")
        title = get_field("title")
        if _is_blank_or_total_row([isbn_raw, title, get_field("author"), get_field("publisher")]):
            continue
        isbn_norm = normalize_isbn(isbn_raw)
        isbn_status = get_isbn_status(isbn_raw)

        award_item = get_field("award_item")

        list_price = _to_float(get_field("list_price"))
        purchase_price = _to_float(get_field("purchase_price"))
        if purchase_price is None and list_price is not None:
            formula = pp_formula_map.get(enum_idx)
            if formula:
                purchase_price = _resolve_formula_purchase_price(formula, list_price)

        book = {
            "title": title,
            "list_price": list_price,
            "purchase_price": purchase_price,
            "author": get_field("author"),
            "publisher": get_field("publisher"),
            "award_item": award_item,
            "eligibility_label": get_field("eligibility_label"),
            "recommendation_source": get_field("recommendation_source"),
        }
        completeness = compute_completeness(book, project_type=proj_type)

        conn.execute(
            "INSERT INTO vendor_books"
            "(batch_id, award_item, vendor_seq, title, author, isbn, isbn_normalized, "
            "publish_date, list_price, purchase_price, publisher, age_range, "
            "isbn_status, completeness_status, extra_fields, source_row_number, raw_row, "
            "category, book_type, policy_topic, summary, source_url, recommendation_source, "
            "eligibility_label, classification_number, award_notes) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                batch_id,
                award_item,
                get_field("vendor_seq"),
                book["title"],
                book["author"],
                isbn_raw,
                isbn_norm,
                get_field("publish_date"),
                book["list_price"],
                book["purchase_price"],
                book["publisher"],
                get_field("age_range"),
                isbn_status,
                completeness,
                None,
                None,
                json.dumps(raw, ensure_ascii=False),
                get_field("category"),
                get_field("book_type"),
                get_field("policy_topic"),
                get_field("summary"),
                get_field("source_url"),
                get_field("recommendation_source"),
                get_field("eligibility_label"),
                get_field("classification_number"),
                get_field("award_notes"),
            ),
        )
        records_inserted += 1

    conn.execute(
        "UPDATE import_batches SET record_count = ? WHERE id = ?",
        (records_inserted, batch_id),
    )
    conn.commit()
    conn.close()

    result = {
        "batch_id": batch_id,
        "record_count": records_inserted,
        "unmapped_fields": unmapped,
        "column_mapping": mapping,
    }
    try:
        result["match_stats"] = run_match(project_id)
    except Exception as e:
        result["match_rerun_error"] = str(e)
    return result


def _to_float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(str(v).replace(",", ""))
    except (ValueError, TypeError):
        return None


def _parse_formula_multiplier(formula: str) -> float | None:
    """Extract multiplier from patterns like *75% → 0.75."""
    m = re.search(r'\*\s*(\d+(?:\.\d+)?)\s*%', formula)
    if m:
        return float(m.group(1)) / 100
    return None


def _resolve_formula_purchase_price(formula: str, list_price: float | None) -> float | None:
    """
    Compute purchase_price from a formula string and list_price.
    Supports =ROUND(ref*N%,0) and =ref*N%. Returns None if pattern unrecognised.
    """
    if list_price is None:
        return None
    multiplier = _parse_formula_multiplier(formula)
    if multiplier is None:
        return None
    if re.search(r'ROUND\s*\(', formula, re.IGNORECASE):
        return float(math.floor(list_price * multiplier + 0.5))
    return list_price * multiplier


def _build_formula_fallback(
    file_bytes: bytes, sheet, header_row: int, pp_src_col: str
) -> dict:
    """
    Return {data_row_0based: formula_string} for the purchase_price column.
    data_row_0based=0 is the first data row (DataFrame index 0).
    Reads openpyxl with data_only=False to capture formula strings.
    Silently returns {} on any error.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=False, read_only=True)
        ws = wb[sheet] if isinstance(sheet, str) else list(wb.worksheets)[int(sheet) if sheet else 0]

        # Locate the purchase_price column by header name (Excel rows are 1-based)
        excel_header_row = header_row + 1
        pp_col_idx = None
        for header_cells in ws.iter_rows(min_row=excel_header_row, max_row=excel_header_row):
            for cell in header_cells:
                if cell.value and str(cell.value).strip() == pp_src_col:
                    pp_col_idx = cell.column
                    break

        if pp_col_idx is None:
            wb.close()
            return {}

        # Data rows start at header_row + 2 (Excel 1-based)
        # data_row_0based = cell.row - (header_row + 2), aligns with DataFrame enum_idx
        excel_data_start = header_row + 2
        formula_map: dict = {}
        for row in ws.iter_rows(min_row=excel_data_start, min_col=pp_col_idx, max_col=pp_col_idx):
            for cell in row:
                val = cell.value
                if val and isinstance(val, str) and val.startswith('='):
                    formula_map[cell.row - excel_data_start] = val

        wb.close()
        return formula_map
    except Exception:
        return {}


def _clear_library_holdings(conn) -> None:
    """
    Remove all library_holdings source data.
    Deletion order: book_matches (holding refs) → library_holdings → import_batches.
    Does NOT commit — caller owns the transaction.
    """
    conn.execute("DELETE FROM book_matches WHERE holding_id IS NOT NULL")
    conn.execute("DELETE FROM library_holdings")
    batch_ids = [
        r[0]
        for r in conn.execute(
            "SELECT id FROM import_batches WHERE batch_type='library_holdings'"
        ).fetchall()
    ]
    if batch_ids:
        ph = ",".join("?" * len(batch_ids))
        conn.execute(f"DELETE FROM import_batches WHERE id IN ({ph})", batch_ids)


def _clear_vendor_books_for_project(conn, project_id: int) -> None:
    """
    Remove vendor_books-related source data for a project before re-import.
    Deletion order: book_matches → vendor_books → import_batches.
    selection_items are NOT deleted — they carry self-contained snapshots.
    Only touches batch_type='vendor_books'; does NOT commit — caller owns the transaction.
    """
    old_batch_ids = [
        r[0]
        for r in conn.execute(
            "SELECT id FROM import_batches WHERE project_id=? AND batch_type='vendor_books'",
            (project_id,),
        ).fetchall()
    ]
    if not old_batch_ids:
        return

    ph = ",".join("?" * len(old_batch_ids))

    conn.execute(
        f"DELETE FROM book_matches WHERE vendor_book_id IN "
        f"(SELECT id FROM vendor_books WHERE batch_id IN ({ph}))",
        old_batch_ids,
    )
    conn.execute(
        f"DELETE FROM vendor_books WHERE batch_id IN ({ph})",
        old_batch_ids,
    )
    conn.execute(
        f"DELETE FROM import_batches WHERE id IN ({ph})",
        old_batch_ids,
    )


def clear_library_holdings(user_id: int) -> dict:
    conn = get_connection()
    try:
        holdings_count = conn.execute("SELECT COUNT(*) FROM library_holdings").fetchone()[0]
        matches_count = conn.execute(
            "SELECT COUNT(*) FROM book_matches WHERE holding_id IS NOT NULL"
        ).fetchone()[0]
        batches_count = conn.execute(
            "SELECT COUNT(*) FROM import_batches WHERE batch_type='library_holdings'"
        ).fetchone()[0]
        _clear_library_holdings(conn)
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return {
        "deleted_holdings": holdings_count,
        "deleted_matches": matches_count,
        "deleted_batches": batches_count,
    }


def clear_vendor_books(project_id: int, user_id: int) -> dict:
    conn = get_connection()
    try:
        project = conn.execute(
            "SELECT id FROM procurement_projects WHERE id=?", (project_id,)
        ).fetchone()
        if project is None:
            raise ValueError(f"project_id {project_id} 不存在")

        old_batch_ids = [
            r[0]
            for r in conn.execute(
                "SELECT id FROM import_batches WHERE project_id=? AND batch_type='vendor_books'",
                (project_id,),
            ).fetchall()
        ]

        if old_batch_ids:
            ph = ",".join("?" * len(old_batch_ids))
            vendor_books_count = conn.execute(
                f"SELECT COUNT(*) FROM vendor_books WHERE batch_id IN ({ph})",
                old_batch_ids,
            ).fetchone()[0]
            matches_count = conn.execute(
                f"SELECT COUNT(*) FROM book_matches WHERE vendor_book_id IN "
                f"(SELECT id FROM vendor_books WHERE batch_id IN ({ph}))",
                old_batch_ids,
            ).fetchone()[0]
        else:
            vendor_books_count = 0
            matches_count = 0

        batches_count = len(old_batch_ids)
        preserved_count = conn.execute(
            "SELECT COUNT(*) FROM selection_items WHERE project_id=?", (project_id,)
        ).fetchone()[0]

        _clear_vendor_books_for_project(conn, project_id)
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return {
        "project_id": project_id,
        "deleted_vendor_books": vendor_books_count,
        "deleted_matches": matches_count,
        "deleted_batches": batches_count,
        "preserved_selection_items": preserved_count,
    }


def _read_library_sheet(
    file_bytes: bytes, engine: str, sheet_name
) -> tuple | None:
    """
    Detect header and match columns for one sheet.
    Returns (df, mapping, unmapped) if the sheet has isbn or title mapping, else None.
    """
    df = _read_excel_with_detected_header(
        file_bytes, engine, LIBRARY_COLUMN_HINTS, sheet_name=sheet_name
    )
    df.columns = [str(c).strip() for c in df.columns]
    mapping, unmapped = _match_columns(list(df.columns), LIBRARY_COLUMN_HINTS)
    if "isbn" not in mapping.values() and "title" not in mapping.values():
        return None
    return df, mapping, unmapped


def _insert_library_holding_rows(conn, batch_id: int, df, mapping: dict) -> int:
    """Insert rows from one sheet into library_holdings. Returns count inserted."""
    reverse_map = {v: k for k, v in mapping.items()}
    records_inserted = 0

    for _, row in df.iterrows():
        raw = {col: (None if pd.isna(row[col]) else str(row[col]).strip()) for col in df.columns}

        def get_field(field: str):
            src = reverse_map.get(field)
            if src:
                v = raw.get(src)
                return v if v else None
            return None

        isbn_raw = get_field("isbn")
        title = get_field("title")
        if _is_blank_or_total_row([isbn_raw, title, get_field("author"), get_field("publisher")]):
            continue
        isbn_norm = normalize_isbn(isbn_raw)
        isbn_status = get_isbn_status(isbn_raw)

        conn.execute(
            "INSERT INTO library_holdings"
            "(batch_id, isbn, isbn_normalized, title, author, publisher, publish_year, "
            "price, library_record_id, isbn_status, raw_row) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                batch_id,
                isbn_raw,
                isbn_norm,
                title,
                get_field("author"),
                get_field("publisher"),
                get_field("publish_year"),
                _to_float(get_field("price")),
                get_field("library_record_id"),
                isbn_status,
                json.dumps(raw, ensure_ascii=False),
            ),
        )
        records_inserted += 1

    return records_inserted
